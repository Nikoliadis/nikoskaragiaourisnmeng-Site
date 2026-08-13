"""Import the links from ΙΣΤΟΣΕΛΙΔΑ.xlsx into the menu categories.

The workbook is the teacher's own map of the site: one sheet per section, a
column of subcategory headings, and the links that belong under each. Reading
it directly means he keeps editing the spreadsheet he already maintains and we
re-run this, rather than anyone retyping forty links into the admin.

    python manage.py import_links                 # add and update
    python manage.py import_links --prune         # also drop links no longer listed
    python manage.py import_links --file other.xlsx
"""

import re
from pathlib import Path
from urllib.parse import unquote

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from content.models import Category, Link, Section

# Sheet name -> the Section it fills.
SHEETS = {
    "ΠΑΝΕΛΛΑΔΙΚΕΣ ΕΞΕΤΑΣΕΙΣ": Section.PANELLINIES,
    "ΒΙΒΛΙΑ ΕΠΑΛ": Section.EBOOKS,
    "ΕΠΑΓΓΕΛΜΑΤΙΚΑ ΛΥΚΕΙΑ": Section.ASKISEIS,
    "ΣΧΟΛΕΣ ΕΠΑΓΓΕΛΜ. ΚΑΤΑΡΤΙΣΗΣ": Section.XRISIMA,
    "ΤΡΙΤΟΒΑΘΜΙΑ ΕΚΠΑΙΔΕΥΣΗ": Section.TRITOVATHMIA,
}

# The workbook's headings do not always match the category names in the menu.
HEADING_TO_CATEGORY = {
    "ΘΕΜΑΤΑ ΚΑΙ ΑΠΑΝΤΗΣΕΙΣ": "ΘΕΜΑΤΑ - ΑΠΑΝΤΗΣΕΙΣ",
    "ΚΑΤΑΛΟΓΟΣ ΕΡΩΤΗΣΕΩΝ ΠΙΣΤΟΠΟΙΗΣΗΣ": "ΚΑΤΑΛΟΓΟΣ ΕΡΩΤΗΣΕΩΝ ΠΙΣΤΟΠΟΙΗΣΗΣ",
}


def clean(value) -> str:
    return " ".join(str(value).split()) if value else ""


# Boilerplate in the ministry's e-book file names: the class, the edition and
# "Βιβλίο Μαθητή" repeated on every single one.
FILENAME_NOISE = re.compile(
    r"(?:^|[-_])(?:\d{2}-\d{4}-\d{2}|V\d|[ABG](?:-[ABG])*-EPAL|EPAL|Vivlio-Mathiti"
    r"|Vivlio|Mathiti|pdf|e[-_]?j\d+)(?=[-_]|$)",
    re.IGNORECASE,
)

# Those file names are Greeklish. Rather than guess at a general
# transliteration, the words that actually occur are spelled out.
GREEKLISH = {
    "stoicheia": "Στοιχεία", "michanon": "Μηχανών", "michanes": "Μηχανές",
    "schedio": "Σχέδιο", "schediasmou": "Σχεδιασμού", "kentrikon": "Κεντρικών",
    "thermanseon": "Θερμάνσεων", "psyksi": "Ψύξη", "klimatismos": "Κλιματισμός",
    "esoterikis": "Εσωτερικής", "kausis": "Καύσης", "kinitires": "Κινητήρες",
    "aeroskafon": "Αεροσκαφών", "ilektrotechnia": "Ηλεκτροτεχνία",
    "ilektrikes": "Ηλεκτρικές", "mathimatika": "Μαθηματικά",
    "statistikis": "Στατιστικής", "nautiliaki": "Ναυτιλιακή",
    "politiki": "Πολιτική", "kai": "και",
}


def readable_from_url(url: str) -> str:
    """Turn a file name into something a student can recognise."""
    tail = [part for part in unquote(url).rstrip("/").split("/") if part]
    name = tail[-1] if tail else url
    name = re.sub(r"\.(pdf|docx?|html?)$", "", name, flags=re.IGNORECASE)
    name = FILENAME_NOISE.sub(" ", name)
    words = [w for w in re.split(r"[-_\s]+", name) if w and not w.isdigit()]
    return " ".join(GREEKLISH.get(w.lower(), w) for w in words).strip()


def tidy_title(title: str, url: str) -> str:
    """Make the spreadsheet's labels read like titles rather than search results."""
    title = unquote(title)
    for tail in (" -Panellinies.net", " – Isalos.net", " | Τράπεζα Θεμάτων",
                 " | eduguide"):
        title = title.replace(tail, "")
    title = title.replace("Επαγγελματικό Λύκειο | Γ' ΤΑΞΗ | ", "")
    title = title.strip(" -–|")

    # A label that is really a URL, a file name or a catalogue code tells a
    # student nothing — rebuild it from the address instead.
    looks_like_a_filename = (
        not title
        or title.startswith(("http", "ebooks.edu.gr", "e_j", "e j"))
        or re.search(r"\.(pdf|docx?)$", title, re.IGNORECASE)
    )
    if looks_like_a_filename:
        title = readable_from_url(url) or title
    return title[:250]


class Command(BaseCommand):
    help = "Import the links from ΙΣΤΟΣΕΛΙΔΑ.xlsx into the menu categories."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=None, help="Path to the workbook.")
        parser.add_argument(
            "--prune", action="store_true",
            help="Remove links that are no longer in the workbook.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Report what would change without touching the database.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - depends on environment
            raise CommandError("openpyxl is required: pip install openpyxl") from exc

        path = Path(options["file"] or Path(__file__).resolve().parents[4] / "ΙΣΤΟΣΕΛΙΔΑ.xlsx")
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True)
        kept, created, updated, skipped = set(), 0, 0, []

        for sheet_name, section in SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                skipped.append(f"sheet «{sheet_name}» missing")
                continue

            sheet = workbook[sheet_name]
            # `parent` is whatever column B last named; `category` may be one of
            # its children. Keeping them apart matters: the row after
            # ΠΑΝΕΠΙΣΤΗΜΙΑ › ΠΡΟΠΤΥΧΙΑΚΑ names ΜΕΤΑΠΤΥΧΙΑΚΑ, which is a sibling
            # of ΠΡΟΠΤΥΧΙΑΚΑ, not a child of it.
            parent = category = None
            for row in sheet.iter_rows():
                heading = clean(row[1].value) if len(row) > 1 else ""
                if heading:
                    found = self._find_category(section, heading)
                    if found is None:
                        skipped.append(f"{sheet_name}: no category for «{heading}»")
                    else:
                        parent = category = found

                # Column C can name a deeper category, but only when it is a
                # heading rather than a link.
                if len(row) > 2 and row[2].value and not row[2].hyperlink:
                    deeper = self._find_category(section, clean(row[2].value), parent=parent)
                    category = deeper or category

                if category is None:
                    continue

                # The hyperlink is not always in the same column — the academies
                # keep theirs in C and the university pages theirs in D.
                cell = next((c for c in row if c.hyperlink and
                             str(c.hyperlink.target).startswith("http")), None)
                if cell is None:
                    continue

                url = cell.hyperlink.target
                title = tidy_title(clean(cell.value), url)
                # Anything else on the row is detail about this entry — the
                # academies list address, telephone and e-mail beside the link.
                # The sheet's first row carries the section title and the column
                # headings, which are labels rather than facts about this entry.
                is_header_row = bool(clean(row[0].value)) if row else False
                extras = [] if is_header_row else [
                    clean(c.value) for c in row
                    if c is not cell and c.value and not c.hyperlink
                ]
                description = " · ".join(e for e in extras if e and e != heading)[:500]

                link, was_created = Link.objects.get_or_create(
                    category=category, url=url,
                    defaults={"title": title, "description": description, "order": cell.row},
                )
                kept.add(link.pk)
                if was_created:
                    created += 1
                    self.stdout.write(f"  + [{category.name}] {title}")
                elif (link.title, link.description) != (title, description) or not link.is_active:
                    link.title, link.description, link.is_active = title, description, True
                    link.save(update_fields=["title", "description", "is_active"])
                    updated += 1

        texts = self._import_text_blocks(workbook)

        self.stdout.write(self.style.SUCCESS(
            f"\n{created} new, {updated} updated, {len(kept)} links in total; "
            f"{texts} categories given text."
        ))
        for note in skipped:
            self.stdout.write(self.style.WARNING(f"  ! {note}"))

        if options["prune"]:
            gone = Link.objects.exclude(pk__in=kept)
            self.stdout.write(f"Removing {gone.count()} links no longer in the workbook.")
            gone.delete()

        if options["dry_run"]:
            transaction.set_rollback(True)
            self.stdout.write(self.style.WARNING("Dry run — nothing was saved."))

    def _import_text_blocks(self, workbook):
        """Carry over the parts of the workbook that are prose, not links.

        The subject list for the school exams is three levels of headings and
        course names with no URL anywhere; it belongs as text on its category.
        """
        sheet = workbook["ΕΠΑΓΓΕΛΜΑΤΙΚΑ ΛΥΚΕΙΑ"]
        category = Category.objects.filter(
            section=Section.ASKISEIS, name__icontains="ΕΝΔΟΣΧΟΛΙΚ"
        ).first()
        if category is None:
            return 0

        html, collecting = [], False
        for row in sheet.iter_rows():
            heading = clean(row[1].value) if len(row) > 1 else ""
            if heading:
                # The block runs from its own heading to the next one.
                collecting = "ΕΝΔΟΣΧΟΛΙΚ" in heading
            if not collecting or len(row) < 3:
                continue
            text = clean(row[2].value)
            if not text or row[2].hyperlink:
                continue

            if text.startswith("ΤΟΜΕΑΣ"):
                html.append(f"<h3>{text}</h3>")
            elif text[:2] in ("α.", "β.", "γ.", "δ.", "ε."):
                html.append(f"<p><strong>{text}</strong></p>")
                html.append("<ul>")
            else:
                html.append(f"<li>{text}</li>")

        if not html:
            return 0

        # Close every list that was opened.
        body = "\n".join(html).replace("<ul>\n<h3>", "</ul>\n<h3>")
        body = body.replace("<ul>\n<p>", "</ul>\n<p>") + "</ul>"
        body = body.replace("<ul>\n</ul>", "")

        if category.description != body:
            category.description = body
            category.save(update_fields=["description"])
            self.stdout.write(f"  ✎ [{category.name}] κατάλογος μαθημάτων")
            return 1
        return 0

    def _find_category(self, section, heading, parent=None):
        """Match a workbook heading to a menu category, tolerating case.

        With a parent given, only its own children are considered, so a word
        like ΠΡΟΠΤΥΧΙΑΚΑ resolves under ΠΑΝΕΠΙΣΤΗΜΙΑ rather than anywhere.
        """
        name = HEADING_TO_CATEGORY.get(heading, heading)
        queryset = Category.objects.filter(section=section)
        if parent is not None:
            queryset = queryset.filter(parent=parent)
        return (
            queryset.filter(name__iexact=name).first()
            or queryset.filter(name__icontains=name).first()
        )
